"""
xps_txt_to_excel.py

Converts XPS .txt files exported by the MATLAB fixXPS pipeline into a
single .xlsx workbook, one sheet per file.

Usage
-----
    python xps_txt_to_excel.py

Edit the CONFIG section below to match your data path and file lists.
"""

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIG
# =============================================================================
DATA_PATH = r'./data/2026-06-26/averaged_data'

LEGENDS = ['Underfocused', 'Focused']

SURVEY_FILES = [
    'aspirin_test_underfocused_Survey_2026-06-26__19h46m59s_alternative.txt',
    'aspirin_test_focused_Survey_2026-06-26__19h28m31s_alternative.txt',
]
C1S_FILES = [
    'aspirin_test_underfocused_C_1s_2026-06-26__20h15m53s_average.txt',
    'aspirin_test_focused_C_1s_2026-06-26__18h52m51s_average.txt',
]
O1S_FILES = [
    'aspirin_test_underfocused_O_1s_2026-06-26__19h56m37s_average.txt',
    'aspirin_test_focused_O_1s_2026-06-26__19h04m33s_average.txt',
]

OUTPUT_FILE = os.path.join(DATA_PATH, 'xps_data.xlsx')

PHOTON_ENERGY = 1486.6  # Al Ka, eV

# =============================================================================
# STYLING HELPERS
# =============================================================================
HEADER_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADER_FILL    = PatternFill('solid', start_color='2F5496')   # dark blue
DATA_FONT      = Font(name='Arial', size=10)
TITLE_FONT     = Font(name='Arial', bold=True, size=12)
CENTER         = Alignment(horizontal='center', vertical='center')
LEFT           = Alignment(horizontal='left',   vertical='center')
THIN           = Side(style='thin', color='AAAAAA')
THIN_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT_FLOAT  = '0.000000'
NUM_FMT_SCI    = '0.00E+00'


def style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def style_data_row(ws, row, n_cols):
    fill = PatternFill('solid', start_color='EEF2FF') if row % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = DATA_FONT
        cell.alignment = CENTER
        cell.fill      = fill
        cell.border    = THIN_BORDER
        # choose number format by column: BE and counts use float, rest scientific
        cell.number_format = NUM_FMT_FLOAT if col <= 2 else NUM_FMT_SCI


def write_sheet(wb, sheet_name, df, legend, spectrum_type):
    ws = wb.create_sheet(title=sheet_name)

    # --- title row ----------------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1,
                         value=f"{legend}  —  {spectrum_type} XPS spectrum")
    title_cell.font      = TITLE_FONT
    title_cell.alignment = LEFT
    title_cell.fill      = PatternFill('solid', start_color='D9E1F2')

    # --- header row ---------------------------------------------------------
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=2, column=col_idx, value=col_name)
    style_header_row(ws, row=2, n_cols=len(df.columns))

    # --- data rows ----------------------------------------------------------
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=float(value))
        style_data_row(ws, row=row_idx, n_cols=len(df.columns))

    # --- column widths ------------------------------------------------------
    col_widths = {'BE_eV': 12, 'KE_eV': 12, 'counts': 14}
    default_width = 18
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, default_width)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = 'A3'   # freeze title + header


# =============================================================================
# MAIN
# =============================================================================
def main():
    all_groups = [
        ('Survey', SURVEY_FILES),
        ('C 1s',   C1S_FILES),
        ('O 1s',   O1S_FILES),
    ]

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    for kk, legend in enumerate(LEGENDS):
        for spectrum_type, file_list in all_groups:
            fpath = os.path.join(DATA_PATH, file_list[kk])

            if not os.path.exists(fpath):
                print(f'  [skip] not found: {fpath}')
                continue

            df = pd.read_csv(fpath, sep=r'\s+', engine='python')

            # compute binding energy from kinetic energy and insert as the
            # first column, so it's exported alongside the raw data
            if 'KE_eV' in df.columns:
                df.insert(0, 'BE_eV', PHOTON_ENERGY - df['KE_eV'])
            else:
                print(f'  [warn] no KE_eV column found in {fpath}; BE_eV not added')

            # sheet names must be ≤31 chars, no special characters
            sheet_name = f"{legend[:10]}_{spectrum_type.replace(' ', '')}"

            write_sheet(wb, sheet_name, df, legend, spectrum_type)
            print(f'  [ok]   {sheet_name}  ({len(df)} rows, {len(df.columns)} cols)')

    wb.save(OUTPUT_FILE)
    print(f'\nSaved: {OUTPUT_FILE}')


if __name__ == '__main__':
    main()